# Interface principal da ferramenta de pesquisa APIB — Sprint 2
# Desenvolvida para mestranda pesquisadora com foco em acessibilidade e clareza

import io
import sqlite3
from datetime import datetime

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import database
from collector import run_collector
from config import DB_PATH, DEFAULT_KEYWORDS

# ---------------------------------------------------------------------------
# Configuração da página (deve ser a primeira chamada Streamlit)
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Pesquisa APIB",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# CSS customizado
# ---------------------------------------------------------------------------
st.markdown("""
<style>
/* Fonte base maior para melhor legibilidade */
html, body, [class*="css"] {
    font-size: 16px !important;
}

/* Reduz padding excessivo do container principal */
.block-container {
    padding-top: 1.5rem !important;
    padding-bottom: 2rem !important;
}

/* Cards de post com borda sutil */
.post-card {
    background-color: #2a2a2a;
    border: 1px solid #3a3a3a;
    border-radius: 8px;
    padding: 1rem 1.2rem;
    margin-bottom: 1rem;
}

/* Botões com padding generoso */
.stButton > button {
    padding: 0.5rem 1.4rem !important;
    font-size: 16px !important;
}

/* Espaçamento entre seções */
.section-gap {
    margin-top: 1.5rem;
}

/* Badge de anotado */
.badge-annotated {
    color: #4caf50;
    font-weight: 600;
}
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Funções auxiliares de banco de dados (com cache de leitura)
# ---------------------------------------------------------------------------

@st.cache_data(ttl=60)
def cached_get_all_posts():
    """Retorna todos os posts com cache de 60 segundos."""
    return database.get_all_posts()


@st.cache_data(ttl=60)
def cached_get_last_collection_date():
    """Retorna a data da última coleta ou None."""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(collected_at) FROM posts")
    row = cursor.fetchone()
    conn.close()
    return row[0] if row and row[0] else None


@st.cache_data(ttl=60)
def cached_get_all_annotations():
    """Retorna todas as anotações como dicionário {shortcode: annotation_dict}."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT post_shortcode, annotation, category, relevant FROM annotations")
    rows = {r["post_shortcode"]: dict(r) for r in cursor.fetchall()}
    conn.close()
    return rows


@st.cache_data(ttl=60)
def cached_dashboard_data():
    """Retorna dados agregados para o dashboard."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) AS total FROM posts")
    total_posts = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM annotations")
    total_annotated = cursor.fetchone()["total"]

    cursor.execute("SELECT COUNT(*) AS total FROM annotations WHERE relevant = 1")
    total_relevant = cursor.fetchone()["total"]

    cursor.execute("SELECT MIN(date) AS min_d, MAX(date) AS max_d FROM posts")
    dates = cursor.fetchone()
    min_date = dates["min_d"]
    max_date = dates["max_d"]

    # Posts por ano
    cursor.execute("""
        SELECT strftime('%Y', date) AS year, COUNT(*) AS count
        FROM posts GROUP BY year ORDER BY year
    """)
    posts_by_year = [dict(r) for r in cursor.fetchall()]

    # Posts por mês
    cursor.execute("""
        SELECT strftime('%Y-%m', date) AS month, COUNT(*) AS count
        FROM posts GROUP BY month ORDER BY month
    """)
    posts_by_month = [dict(r) for r in cursor.fetchall()]

    # Posts por categoria
    cursor.execute("""
        SELECT category, COUNT(*) AS count
        FROM annotations WHERE category IS NOT NULL AND category != ''
        GROUP BY category ORDER BY count DESC
    """)
    posts_by_category = [dict(r) for r in cursor.fetchall()]

    # Top 10 por likes
    cursor.execute("""
        SELECT shortcode, url, date, likes,
               SUBSTR(caption, 1, 80) AS caption_preview
        FROM posts ORDER BY likes DESC LIMIT 10
    """)
    top_likes = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return {
        "total_posts": total_posts,
        "total_annotated": total_annotated,
        "total_relevant": total_relevant,
        "min_date": min_date,
        "max_date": max_date,
        "posts_by_year": posts_by_year,
        "posts_by_month": posts_by_month,
        "posts_by_category": posts_by_category,
        "top_likes": top_likes,
    }


def filter_posts(posts, keyword, year_start, year_end):
    """Filtra lista de posts por palavra-chave e intervalo de anos."""
    results = []
    for p in posts:
        post_year = int(p["date"][:4]) if p.get("date") else 0
        if post_year < year_start or post_year > year_end:
            continue
        if keyword:
            if keyword.lower() not in (p.get("caption") or "").lower():
                continue
        results.append(p)
    return results


def format_date(date_str):
    """Converte YYYY-MM-DD para DD/MM/AAAA."""
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return date_str or "—"


def build_excel(posts_df):
    """
    Gera um arquivo Excel em memória com formatação.
    Retorna bytes do arquivo .xlsx.
    """
    output = io.BytesIO()
    wb = Workbook()

    # Aba principal
    ws = wb.active
    ws.title = "Posts"

    header_fill = PatternFill(start_color="2a2a2a", end_color="2a2a2a", fill_type="solid")
    header_font = Font(bold=True, color="E0E0E0")

    columns = list(posts_df.columns)
    ws.append(columns)

    # Formata cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Dados
    for _, row in posts_df.iterrows():
        ws.append(list(row))

    # Largura das colunas baseada no cabeçalho e nos primeiros 100 valores
    for col in ws.columns:
        sample = [col[0]] + list(col[1:101])
        max_len = max((len(str(cell.value or "")) for cell in sample), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    # Aba Dashboard com resumo
    ws2 = wb.create_sheet("Dashboard")
    ws2.append(["Resumo", "Valor"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2.append(["Total de posts", len(posts_df)])

    if "data" in posts_df.columns:
        valid_dates = posts_df["data"].dropna().astype(str)
        valid_dates = valid_dates[valid_dates.str.len() >= 4]
        if not valid_dates.empty:
            by_year = valid_dates.str[:4].value_counts().sort_index().reset_index()
            by_year.columns = ["ano", "total"]
            ws2.append([])
            ws2.append(["Ano", "Posts"])
            for _, r in by_year.iterrows():
                ws2.append([r["ano"], r["total"]])

    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Inicialização do banco
# ---------------------------------------------------------------------------
database.init_db()

# ---------------------------------------------------------------------------
# Sidebar — navegação
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown("## 🌿 Pesquisa APIB")
    st.caption("Ferramenta de Análise")
    st.divider()

    secao = st.radio(
        "Navegação",
        options=[
            "📥 Coleta",
            "🔍 Explorar Posts",
            "📝 Anotações",
            "📊 Dashboard",
            "💾 Exportar",
        ],
        label_visibility="collapsed",
    )

# ===========================================================================
# Seção 1 — Coleta
# ===========================================================================
if secao == "📥 Coleta":
    st.title("📥 Coleta de Publicações")

    st.write(
        "Clique no botão abaixo para iniciar a coleta de posts da APIB no Instagram. "
        "Isso pode levar alguns minutos."
    )

    # Informação sobre o banco atual
    all_posts = cached_get_all_posts()
    if all_posts:
        last_collected = cached_get_last_collection_date()
        st.info(
            f"Você já tem **{len(all_posts)} posts** coletados. "
            "A coleta vai adicionar apenas posts novos."
        )
        if last_collected:
            try:
                dt = datetime.fromisoformat(last_collected)
                st.caption(f"Última coleta: {dt.strftime('%d/%m/%Y às %H:%M')}")
            except Exception:
                pass

    if st.button("Iniciar Coleta", type="primary"):
        progress_bar = st.progress(0, text="Iniciando coleta...")
        status_text = st.empty()
        result_container = st.empty()

        collected_count = [0]

        def on_progress(count, date_str):
            collected_count[0] = count
            # Progresso indeterminado: avança lentamente até 0.95 (sem saber o total)
            progress_val = min(0.95, count / (count + 20))
            formatted = format_date(date_str)
            progress_bar.progress(progress_val, text="Coletando posts...")
            status_text.text(f"Posts coletados até agora: {count} | Post de {formatted}")

        with st.spinner("Coletando posts... Por favor, aguarde."):
            total = run_collector(progress_callback=on_progress)

        progress_bar.progress(1.0, text="Concluído!")
        status_text.empty()
        # Invalida caches de leitura após nova coleta
        cached_get_all_posts.clear()
        cached_get_last_collection_date.clear()
        st.success(f"Coleta finalizada! {total} posts coletados.")

# ===========================================================================
# Seção 2 — Explorar Posts
# ===========================================================================
elif secao == "🔍 Explorar Posts":
    st.title("🔍 Explorar Publicações")

    all_posts = cached_get_all_posts()

    if not all_posts:
        st.warning(
            "Nenhum post coletado ainda. Vá para a seção **Coleta** para começar."
        )
    else:
        all_annotations = cached_get_all_annotations()

        # Filtros no topo
        col1, col2 = st.columns([2, 2])
        with col1:
            keyword_options = [""] + DEFAULT_KEYWORDS + ["Digitar outra..."]
            keyword_select = st.selectbox(
                "Palavra-chave predefinida",
                options=keyword_options,
                format_func=lambda x: "Selecione ou deixe em branco" if x == "" else x,
            )
            if keyword_select == "Digitar outra...":
                keyword = st.text_input("Digite a palavra-chave")
            else:
                keyword = keyword_select

        with col2:
            year_range = st.slider("Período (ano)", 2019, 2025, (2019, 2025))

        if st.button("Buscar", type="primary"):
            filtered = filter_posts(all_posts, keyword, year_range[0], year_range[1])

            if not filtered:
                st.info(
                    "Nenhum post encontrado para esse filtro. Tente outros termos."
                )
            else:
                st.write(f"**{len(filtered)} publicações encontradas**")
                for post in filtered:
                    shortcode = post.get("shortcode", "")
                    annotation = all_annotations.get(shortcode)
                    annotated_badge = (
                        '<span class="badge-annotated">✅ Anotado</span>'
                        if annotation
                        else ""
                    )

                    st.markdown(
                        f'<div class="post-card">'
                        f'📅 <strong>{format_date(post.get("date", ""))}</strong> &nbsp;|&nbsp; '
                        f'<a href="{post.get("url", "#")}" target="_blank">🔗 Abrir no Instagram</a>'
                        f'&nbsp;&nbsp;{annotated_badge}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    with st.expander("Ver legenda completa"):
                        st.write(post.get("caption") or "*(sem legenda)*")

                    like_col, comment_col = st.columns(2)
                    like_col.write(f"👍 {post.get('likes', 0)} curtidas")
                    comment_col.write(f"💬 {post.get('comments', 0)} comentários")
                    st.divider()

# ===========================================================================
# Seção 3 — Anotações
# ===========================================================================
elif secao == "📝 Anotações":
    st.title("📝 Anotações por Post")

    all_posts = cached_get_all_posts()

    if not all_posts:
        st.warning(
            "Nenhum post coletado ainda. Vá para a seção **Coleta** para começar."
        )
    else:
        all_annotations = cached_get_all_annotations()

        # Filtros
        col1, col2 = st.columns([2, 2])
        with col1:
            keyword_options = [""] + DEFAULT_KEYWORDS + ["Digitar outra..."]
            keyword_select = st.selectbox(
                "Palavra-chave predefinida",
                options=keyword_options,
                format_func=lambda x: "Selecione ou deixe em branco" if x == "" else x,
                key="annot_keyword_select",
            )
            if keyword_select == "Digitar outra...":
                keyword = st.text_input("Digite a palavra-chave", key="annot_keyword_input")
            else:
                keyword = keyword_select

        with col2:
            year_range = st.slider(
                "Período (ano)", 2019, 2025, (2019, 2025), key="annot_year_range"
            )

        if st.button("Buscar", type="primary", key="annot_search"):
            filtered = filter_posts(all_posts, keyword, year_range[0], year_range[1])

            if not filtered:
                st.info(
                    "Nenhum post encontrado para esse filtro. Tente outros termos."
                )
            else:
                st.write(f"**{len(filtered)} publicações encontradas**")

                categories = [
                    "Sem categoria",
                    "Mobilização",
                    "Denúncia",
                    "Conquista",
                    "Discurso político",
                    "Cultura indígena",
                    "Outro",
                ]

                for post in filtered:
                    shortcode = post.get("shortcode", "")
                    existing = all_annotations.get(shortcode) or {}
                    caption = post.get("caption") or ""
                    preview = (caption[:200] + "...") if len(caption) > 200 else caption

                    st.markdown(
                        f"**📅 {format_date(post.get('date', ''))}** — "
                        f"[Abrir no Instagram]({post.get('url', '#')})"
                    )
                    st.caption(preview or "*(sem legenda)*")

                    annot_text = st.text_area(
                        "Sua anotação",
                        value=existing.get("annotation", ""),
                        key=f"annot_text_{shortcode}",
                    )

                    existing_cat = existing.get("category", "Sem categoria") or "Sem categoria"
                    cat_index = categories.index(existing_cat) if existing_cat in categories else 0
                    annot_cat = st.selectbox(
                        "Categoria",
                        categories,
                        index=cat_index,
                        key=f"annot_cat_{shortcode}",
                    )

                    annot_relevant = st.checkbox(
                        "Marcar como relevante para a pesquisa",
                        value=bool(existing.get("relevant", 0)),
                        key=f"annot_rel_{shortcode}",
                    )

                    if st.button("Salvar anotação", key=f"save_{shortcode}"):
                        database.save_annotation(
                            shortcode=shortcode,
                            annotation=annot_text,
                            category=annot_cat,
                            relevant=1 if annot_relevant else 0,
                        )
                        cached_get_all_annotations.clear()
                        cached_dashboard_data.clear()
                        st.success("Anotação salva!")

                    st.divider()

# ===========================================================================
# Seção 4 — Dashboard
# ===========================================================================
elif secao == "📊 Dashboard":
    st.title("📊 Painel de Análise")

    data = cached_dashboard_data()

    if data["total_posts"] == 0:
        st.info(
            "Nenhum dado disponível ainda. Vá para a seção **Coleta** para começar."
        )
    else:
        # Métricas de destaque
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total de posts", data["total_posts"])
        m2.metric("Posts anotados", data["total_annotated"])
        m3.metric("Posts relevantes", data["total_relevant"])

        periodo = "—"
        if data["min_date"] and data["max_date"]:
            periodo = f"{format_date(data['min_date'])} → {format_date(data['max_date'])}"
        m4.metric("Período coberto", periodo)

        st.markdown('<div class="section-gap"></div>', unsafe_allow_html=True)

        # Gráfico 1 — Posts por ano
        with st.container():
            st.subheader("Posts por ano")
            if data["posts_by_year"]:
                df_year = pd.DataFrame(data["posts_by_year"])
                fig1 = px.bar(
                    df_year,
                    x="year",
                    y="count",
                    labels={"year": "Ano", "count": "Posts"},
                    template="plotly_dark",
                    color_discrete_sequence=["#5b8dee"],
                )
                fig1.update_layout(
                    paper_bgcolor="#1a1a1a",
                    plot_bgcolor="#1a1a1a",
                    font_color="#e0e0e0",
                )
                st.plotly_chart(fig1, use_container_width=True)
            else:
                st.info("Dados insuficientes para este gráfico.")

        # Gráfico 2 — Posts por categoria
        with st.container():
            st.subheader("Posts por categoria (anotações)")
            if data["posts_by_category"]:
                df_cat = pd.DataFrame(data["posts_by_category"])
                fig2 = px.bar(
                    df_cat,
                    x="count",
                    y="category",
                    orientation="h",
                    labels={"category": "Categoria", "count": "Posts"},
                    template="plotly_dark",
                    color_discrete_sequence=["#5b8dee"],
                )
                fig2.update_layout(
                    paper_bgcolor="#1a1a1a",
                    plot_bgcolor="#1a1a1a",
                    font_color="#e0e0e0",
                )
                st.plotly_chart(fig2, use_container_width=True)
            else:
                st.info("Nenhuma anotação com categoria encontrada.")

        # Gráfico 3 — Posts por mês (linha)
        with st.container():
            st.subheader("Posts por mês ao longo do tempo")
            if data["posts_by_month"]:
                df_month = pd.DataFrame(data["posts_by_month"])
                fig3 = px.line(
                    df_month,
                    x="month",
                    y="count",
                    labels={"month": "Mês", "count": "Posts"},
                    template="plotly_dark",
                    color_discrete_sequence=["#5b8dee"],
                )
                fig3.update_layout(
                    paper_bgcolor="#1a1a1a",
                    plot_bgcolor="#1a1a1a",
                    font_color="#e0e0e0",
                )
                st.plotly_chart(fig3, use_container_width=True)
            else:
                st.info("Dados insuficientes para este gráfico.")

        # Gráfico 4 — Top 10 posts por likes
        with st.container():
            st.subheader("Top 10 posts com mais curtidas")
            if data["top_likes"]:
                df_top = pd.DataFrame(data["top_likes"])
                df_top["label"] = df_top.apply(
                    lambda r: f"{format_date(r['date'])} — {r['caption_preview']}...", axis=1
                )
                fig4 = px.bar(
                    df_top,
                    x="likes",
                    y="label",
                    orientation="h",
                    labels={"likes": "Curtidas", "label": "Post"},
                    template="plotly_dark",
                    color_discrete_sequence=["#5b8dee"],
                )
                fig4.update_layout(
                    paper_bgcolor="#1a1a1a",
                    plot_bgcolor="#1a1a1a",
                    font_color="#e0e0e0",
                    height=420,
                )
                st.plotly_chart(fig4, use_container_width=True)
            else:
                st.info("Dados insuficientes para este gráfico.")

# ===========================================================================
# Seção 5 — Exportar
# ===========================================================================
elif secao == "💾 Exportar":
    st.title("💾 Exportar Dados")

    st.write(
        "Exporte as publicações coletadas para analisar em outros programas."
    )

    all_posts = cached_get_all_posts()
    all_annotations = cached_get_all_annotations()

    if not all_posts:
        st.warning(
            "Nenhum post coletado ainda. Vá para a seção **Coleta** para começar."
        )
    else:
        # Monta DataFrame completo
        rows = []
        for p in all_posts:
            sc = p.get("shortcode", "")
            ann = all_annotations.get(sc) or {}
            rows.append({
                "data": p.get("date", ""),
                "url": p.get("url", ""),
                "legenda": p.get("caption", ""),
                "likes": p.get("likes", 0),
                "comentários": p.get("comments", 0),
                "tipo": p.get("post_type", ""),
                "anotação": ann.get("annotation", ""),
                "categoria": ann.get("category", ""),
                "relevante": ann.get("relevant", 0),
            })

        df_all = pd.DataFrame(rows)
        df_relevant = df_all[df_all["relevante"] == 1]

        st.divider()

        # Opção 1 — CSV
        st.subheader("Opção 1 — CSV completo")
        csv_bytes = df_all.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
        st.download_button(
            label="⬇️ Baixar CSV",
            data=csv_bytes,
            file_name="apib_posts.csv",
            mime="text/csv",
        )

        st.divider()

        # Opção 2 — Excel formatado
        st.subheader("Opção 2 — Excel formatado")
        excel_bytes = build_excel(df_all)
        st.download_button(
            label="⬇️ Baixar Excel",
            data=excel_bytes,
            file_name="apib_posts.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.divider()

        # Opção 3 — Apenas relevantes
        st.subheader("Opção 3 — Apenas posts relevantes")
        if df_relevant.empty:
            st.info(
                "Nenhum post marcado como relevante ainda. "
                "Use a seção **Anotações** para marcar posts relevantes."
            )
            st.button("⬇️ Baixar apenas os Relevantes (Excel)", disabled=True)
        else:
            excel_rel_bytes = build_excel(df_relevant)
            st.download_button(
                label="⬇️ Baixar apenas os Relevantes (Excel)",
                data=excel_rel_bytes,
                file_name="apib_posts_relevantes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
